from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT_DIR = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT_DIR / "Sushicious_Documents.xlsx"


def _set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _apply_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    side = Side(style="thin", color="9AA0A6")
    border = Border(left=side, right=side, top=side, bottom=side)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = border


def _title(ws, text: str) -> None:
    ws["A1"] = text
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28


def _section(ws, row: int, title: str) -> int:
    ws[f"A{row}"] = title
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill("solid", fgColor="1A73E8")
    ws[f"A{row}"].alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20
    return row + 1


def _kv_table(
    ws,
    start_row: int,
    rows: list[tuple[str, str]],
    key_col: str = "A",
    val_col: str = "B",
    merge_val_to: str = "H",
) -> int:
    r = start_row
    for k, v in rows:
        ws[f"{key_col}{r}"] = k
        ws[f"{key_col}{r}"].font = Font(bold=True)
        ws[f"{key_col}{r}"].alignment = Alignment(vertical="top", wrap_text=True)

        ws[f"{val_col}{r}"] = v
        ws.merge_cells(f"{val_col}{r}:{merge_val_to}{r}")
        ws[f"{val_col}{r}"].alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    _apply_border(ws, start_row, r - 1, 1, 8)
    return r + 1


def _table(
    ws,
    start_row: int,
    headers: list[str],
    data: list[list[str]],
    col_widths: list[float] | None = None,
) -> int:
    header_fill = PatternFill("solid", fgColor="E8F0FE")
    header_font = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    r = start_row + 1
    for row in data:
        for i, v in enumerate(row, start=1):
            ws.cell(row=r, column=i, value=v).alignment = wrap
        r += 1

    _apply_border(ws, start_row, r - 1, 1, len(headers))

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

    return r + 1


def _sheet_requirements(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "要件定義書"
    _set_col_widths(ws, {"A": 22, "B": 90, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18})
    _title(ws, "要件定義書（現状実装ベース）")
    ws["H2"] = f"作成日: {date.today().isoformat()}"
    ws["H2"].alignment = Alignment(horizontal="right")

    r = 3
    r = _section(ws, r, "1. 概要")
    r = _kv_table(
        ws,
        r,
        [
            ("システム名", "Sushicious Games（ブラウザゲームポータル）"),
            ("目的", "インストール不要の寿司テーマのミニゲームを提供し、ランキングやトレンド記事で回遊を促進する。"),
            (
                "提供価値",
                "モバイル最適化・短時間で遊べるゲーム体験／日次更新トレンド記事／世界ランキング（Firestore）＋ローカルランキング（localStorage）。",
            ),
            (
                "対象ユーザ",
                "一般ユーザ（PC/スマートフォン）。ログインなしでプレイ可能。",
            ),
        ],
    )

    r = _section(ws, r, "2. 対象範囲（スコープ）")
    r = _kv_table(
        ws,
        r,
        [
            (
                "公開Web",
                "トップ（index.html）／ゲーム詳細（Web/game.html）／トレンド（Web/trends.html + Web/archives/*）／お問い合わせ（Web/contact.html）／利用規約（Web/terms.html）／プライバシー（Web/privacy.html）／各ゲーム（Web/games/*）。",
            ),
            ("ホスティング", "Firebase Hosting（public: Web, cleanUrls: true）。"),
            ("自動更新", "GitHub Actions（日次）→ TrendsUpdater（Gemini API）→ index.html/トレンドHTML/アーカイブHTML差し替え。"),
            ("ランキング", "Firebase Firestore（各ゲーム別コレクション）＋ localStorage（端末内上位3）。"),
        ],
    )

    r = _section(ws, r, "3. 画面・機能要件（要約）")
    r = _table(
        ws,
        r,
        ["区分", "要件", "補足（現状）"],
        [
            ["共通", "言語切替（JP/EN）", "localStorage: sushicious_lang。data-i18n を置換。"],
            ["トップ", "ゲーム一覧（カード表示）", "site.js の games 配列から動的生成。"],
            ["トップ", "最新トレンドのサマリ表示", "index.html の #trends-summary-container を日次で差し替え。"],
            ["ゲーム詳細", "クエリ id に応じてゲームを埋め込み表示", "iframe src に games[*].sceneUrl を設定。"],
            ["ゲーム詳細", "説明／遊び方／攻略の表示", "translations の descLong/ruleKeys/tipsKey を表示。"],
            ["ゲーム（共通）", "ゲーム開始〜終了の進行", "Canvas/Phaser による実装。親ページへ状態通知するものあり。"],
            ["ゲーム（共通）", "ランキング（ローカル）", "localStorage に上位3件保存。ゲームにより日次/通算が存在。"],
            ["ゲーム（共通）", "ランキング（グローバル）", "Firestore コレクションに score/timestamp/userAgent を保存し上位3件取得。"],
            ["トレンド", "日次トレンド本文＋過去アーカイブへの導線", "Web/trends.html に当日記事、Web/archives/* に過去記事。"],
            ["共通", "規約／プライバシー／問い合わせの提供", "静的ページ。問い合わせはメールアドレス提示。"],
        ],
        col_widths=[14, 52, 42],
    )

    r = _section(ws, r, "4. 非機能要件（現状からの整理）")
    r = _table(
        ws,
        r,
        ["分類", "要件", "補足"],
        [
            ["性能", "モバイルでも快適に動作すること", "軽量な静的配信。画像は lazy load を使用。"],
            ["可用性", "常時アクセス可能であること", "Firebase Hosting の静的配信。"],
            ["保守性", "ゲーム追加が容易であること", "site.js の games 配列へ追加し game.html で共通表示。"],
            ["セキュリティ", "秘密情報をクライアントに埋め込まないこと", "トレンド生成は Tools 側で実施。鍵はCI環境変数想定。"],
            ["プライバシー", "ユーザ識別を最小化すること", "ランキング保存に userAgent を利用。個人情報入力なし。"],
            ["SEO", "OGP/description の適切な設定", "index/trends は meta/OGP を設定。"],
        ],
        col_widths=[12, 54, 42],
    )

    r = _section(ws, r, "5. 制約・前提")
    _kv_table(
        ws,
        r,
        [
            ("前提", "ログインなし。ランキングは匿名スコア登録。"),
            ("制約", "ゲームの実行は iframe 埋め込みのため、ブラウザ制限/追跡防止設定の影響を受ける可能性がある。"),
            ("制約", "トレンド更新は外部API（Gemini）に依存し、生成結果の品質はモデル出力に依存する。"),
        ],
    )

    ws.freeze_panes = "A3"


def _sheet_architecture(wb: Workbook) -> None:
    ws = wb.create_sheet("構成図")
    _set_col_widths(ws, {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18})
    _title(ws, "構成図（論理構成 / 現状）")

    r = 3
    r = _section(ws, r, "1. 論理構成（テキスト図）")
    diagram = "\n".join(
        [
            "利用者ブラウザ（PC/モバイル）",
            "  ├─ 静的配信: Firebase Hosting",
            "  │    ├─ HTML/CSS/JS（index / game / trends / contact / terms / privacy）",
            "  │    └─ 各ゲーム（Canvas 2D / Phaser 3 Matter）",
            "  ├─ ランキング保存/取得: Firebase Firestore",
            "  │    └─ collections: rankings_tap / rankings_catch / rankings_match / rankings_docking",
            "  ├─ 広告/計測: Google AdSense（adsbygoogle.js）, アフィリエイトHTML（A8/Rakuten）",
            "  └─ SNS導線: X（share link）",
            "",
            "運用（コンテンツ更新）",
            "  └─ GitHub Actions（日次）",
            "       └─ Tools/TrendsUpdater/update_trends.py（Gemini API）",
            "            ├─ index.html の trends-summary を差し替え",
            "            └─ Web/trends.html / Web/archives/* を生成・更新",
        ]
    )
    ws["A4"] = diagram
    ws.merge_cells("A4:H24")
    ws["A4"].alignment = Alignment(vertical="top", wrap_text=True)
    ws["A4"].font = Font(name="Consolas")
    _apply_border(ws, 4, 24, 1, 8)

    ws.freeze_panes = "A3"


def _sheet_screen_transition(wb: Workbook) -> None:
    ws = wb.create_sheet("画面遷移図")
    _set_col_widths(ws, {"A": 18, "B": 28, "C": 28, "D": 50, "E": 18, "F": 18, "G": 18, "H": 18})
    _title(ws, "画面遷移図（現状）")

    r = 3
    r = _section(ws, r, "1. 遷移一覧")
    r = _table(
        ws,
        r,
        ["From", "To", "トリガー", "URL/パラメータ例"],
        [
            ["トップ", "ゲーム詳細", "ゲームカード選択", "Web/game?id=sushi-tap（id は games[*].id）"],
            ["トップ", "トレンド", "ナビ/ボタン選択", "Web/trends.html"],
            ["トップ", "お問い合わせ", "ナビ/フッタ選択", "Web/contact.html"],
            ["トップ", "利用規約", "フッタ選択", "Web/terms.html"],
            ["トップ", "プライバシー", "フッタ選択", "Web/privacy.html"],
            ["ゲーム詳細", "トップ", "戻るリンク", "../index.html#game（現状は #game 表記）"],
            ["トレンド", "過去アーカイブ", "過去リンク選択", "Web/archives/trend-YYYYMMDD.html"],
            ["トレンド/他", "トップ", "ロゴ選択", "../index.html"],
        ],
        col_widths=[14, 14, 20, 64],
    )

    r = _section(ws, r, "2. 簡易遷移図（テキスト）")
    diag = "\n".join(
        [
            "[トップ/index.html]",
            "  ├─(カード選択)→ [ゲーム詳細/Web/game.html?id=...] ─(戻る)→ [トップ]",
            "  ├─(ナビ)→ [トレンド/Web/trends.html] ─(過去)→ [アーカイブ/Web/archives/*]",
            "  ├─(ナビ)→ [お問い合わせ/Web/contact.html]",
            "  ├─(フッタ)→ [利用規約/Web/terms.html]",
            "  └─(フッタ)→ [プライバシー/Web/privacy.html]",
        ]
    )
    ws["A" + str(r)] = diag
    ws.merge_cells(f"A{r}:H{r+10}")
    ws["A" + str(r)].alignment = Alignment(vertical="top", wrap_text=True)
    ws["A" + str(r)].font = Font(name="Consolas")
    _apply_border(ws, r, r + 10, 1, 8)

    ws.freeze_panes = "A3"


def _sheet_screen_design(wb: Workbook) -> None:
    ws = wb.create_sheet("画面設計書")
    _set_col_widths(ws, {"A": 16, "B": 22, "C": 32, "D": 52, "E": 52, "F": 18, "G": 18, "H": 18})
    _title(ws, "画面設計書（現状実装ベース）")

    r = 3
    r = _section(ws, r, "1. 画面一覧")
    r = _table(
        ws,
        r,
        ["画面ID", "画面名", "パス", "主要要素（ID/クラス）", "主な操作", "備考"],
        [
            [
                "SCR-001",
                "トップ",
                "/index.html",
                "header, #lang-toggle, #game-list, #trends-summary-container, footer",
                "言語切替／ゲームカード選択／トレンド遷移",
                "ゲーム一覧は site.js で動的生成",
            ],
            [
                "SCR-002",
                "ゲーム詳細",
                "/Web/game.html",
                "#game-title, #game-short-desc, #game-iframe, #game-long-desc, #game-rules, #game-tips",
                "言語切替／戻る／ゲームプレイ（iframe内）",
                "query: id で内容切替",
            ],
            [
                "SCR-003",
                "トレンド",
                "/Web/trends.html",
                "#trends-container, .trend-article, .archive",
                "言語切替／過去記事選択",
                "日次でHTML更新",
            ],
            [
                "SCR-004",
                "お問い合わせ",
                "/Web/contact.html",
                "本文（連絡先メール）",
                "言語切替／メールアドレス確認",
                "静的",
            ],
            [
                "SCR-005",
                "利用規約",
                "/Web/terms.html",
                "本文",
                "閲覧のみ",
                "静的",
            ],
            [
                "SCR-006",
                "プライバシー",
                "/Web/privacy.html",
                "本文",
                "閲覧のみ",
                "静的",
            ],
        ],
        col_widths=[12, 16, 22, 44, 44, 22],
    )

    r = _section(ws, r, "2. ゲーム一覧（埋め込み先）")
    _table(
        ws,
        r,
        ["ゲームID", "表示名（JP）", "種別", "埋め込みURL（sceneUrl）", "実装方式", "ランキング"],
        [
            ["sushi-tap", "寿司タップ", "アクション", "games/sushi-tap/Scenes/", "Canvas 2D", "localStorage + Firestore(rankings_tap)"],
            ["sushi-catch", "寿司キャッチ", "キャッチ", "games/sushi-catch/Scenes/", "Canvas 2D", "localStorage + Firestore(rankings_catch)"],
            ["sushi-match", "寿司マッチ", "パズル", "games/sushi-match/Scenes/", "Canvas 2D", "localStorage + Firestore(rankings_match)"],
            ["sushi-docking", "寿司ドッキング", "物理合体パズル", "games/sushi-docking/Scenes/", "Phaser 3 (Matter)", "localStorage + Firestore(rankings_docking)"],
        ],
        col_widths=[14, 18, 14, 28, 18, 32],
    )

    ws.freeze_panes = "A3"


def _sheet_parameters(wb: Workbook) -> None:
    ws = wb.create_sheet("パラメータシート")
    _set_col_widths(ws, {"A": 28, "B": 14, "C": 18, "D": 30, "E": 52, "F": 18, "G": 18, "H": 18})
    _title(ws, "パラメータシート（主要設定/データ項目）")

    r = 3
    r = _section(ws, r, "1. フロント設定/状態")
    r = _table(
        ws,
        r,
        ["パラメータ", "種別", "デフォルト/例", "格納/取得先", "説明"],
        [
            ["sushicious_lang", "localStorage", "ja または en", "localStorage", "言語切替の保持。未設定時は navigator.language を参照。"],
            ["id", "query", "sushi-tap など", "Web/game.html?...", "ゲーム詳細で表示ゲームを特定する識別子。games 配列の id と一致。"],
            ["games[*].id", "定義", "sushi-tap", "Web/Scripts/site.js", "ゲーム識別子。URLクエリに利用。"],
            ["games[*].sceneUrl", "定義", "games/sushi-tap/Scenes/", "Web/Scripts/site.js", "iframe に設定する埋め込み先（ディレクトリ/URL）。"],
            ["games[*].titleKey/descKey/...", "定義", "game_sushi_tap_title 等", "Web/Scripts/site.js", "translations 参照用キー。"],
            ["translations[lang][key]", "辞書", "nav_games 等", "Web/Scripts/site.js", "data-i18n の表示文言およびゲーム説明文を提供。"],
        ],
        col_widths=[26, 12, 18, 24, 60],
    )

    r = _section(ws, r, "2. ランキング（Firestore/ローカル）")
    r = _table(
        ws,
        r,
        ["項目", "種別", "例", "対象", "説明"],
        [
            ["rankings_tap", "Firestore collection", "-", "Sushi Tap", "スコア保存/上位取得。"],
            ["rankings_catch", "Firestore collection", "-", "Sushi Catch", "スコア保存/上位取得。"],
            ["rankings_match", "Firestore collection", "-", "Sushi Match", "スコア保存/上位取得。"],
            ["rankings_docking", "Firestore collection", "-", "SUSHI-Docking", "スコア保存/上位取得。"],
            ["score", "Firestore field", "12345", "全ゲーム", "スコア値。"],
            ["timestamp", "Firestore field", "serverTimestamp()", "全ゲーム", "登録時刻。"],
            ["userAgent", "Firestore field", "Mozilla/5.0 ...", "全ゲーム", "端末/ブラウザ情報（匿名）。"],
            ["local ranking", "localStorage", "上位3件", "全ゲーム", "端末内のランキング。ゲームにより日次/通算が存在。"],
        ],
        col_widths=[22, 18, 18, 18, 54],
    )

    r = _section(ws, r, "3. トレンド更新（運用）")
    _table(
        ws,
        r,
        ["項目", "種別", "例", "場所", "説明"],
        [
            ["GitHub Actions", "ジョブ", "update_trends.yml", ".github/workflows", "日次でトレンド更新を実行し、生成物をコミット。"],
            ["Gemini API", "外部API", "JSON生成", "Tools/TrendsUpdater", "AI/ゲーム/寿司関連のトレンドJSONを生成。"],
            ["index.html trends-summary", "HTML差替", "#trends-summary-container", "index.html", "トップのサマリ本文を日次更新。"],
            ["Web/trends.html", "HTML生成", "当日記事", "Web/trends.html", "最新トレンド記事本文。"],
            ["Web/archives/*", "HTML生成", "trend-YYYYMMDD.html", "Web/archives", "過去トレンド記事アーカイブ。"],
        ],
        col_widths=[22, 14, 22, 20, 54],
    )

    ws.freeze_panes = "A3"


def main() -> None:
    wb = Workbook()
    wb.properties.title = "Sushicious Documents"
    wb.properties.creator = "Doc Generator"

    _sheet_requirements(wb)
    _sheet_architecture(wb)
    _sheet_screen_transition(wb)
    _sheet_screen_design(wb)
    _sheet_parameters(wb)

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    main()

